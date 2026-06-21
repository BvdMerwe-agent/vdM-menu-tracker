#!/usr/bin/env python3
"""
Export the 12-week meal plan + ingredients to a spreadsheet (CSV/XLSX).

Usage:
  python3 export_spreadsheet.py --format xlsx --output vdM_MealPlan.xlsx
  python3 export_spreadsheet.py --format csv --output-dir ./exports/
"""
import sys
import argparse
import csv
from pathlib import Path
from collections import defaultdict

sys.path.insert(0, str(Path(__file__).parent))
from core import get_conn, get_active_plan, get_current_week_number, get_plan_week, get_week_menu

def get_full_plan(plan_id: int) -> list:
    """Return a flat list of all meals across all 12 weeks."""
    conn = get_conn()
    rows = conn.execute(
        """
        SELECT 
            mpw.week_number,
            mpm.day_of_week,
            mpm.meal_type,
            r.id as recipe_id,
            r.name as recipe_name,
            r.cuisine,
            r.prep_time,
            r.cook_time,
            r.servings,
            r.difficulty,
            r.tags
        FROM meal_plan_meals mpm
        JOIN meal_plan_weeks mpw ON mpm.week_id = mpw.id
        LEFT JOIN recipes r ON mpm.recipe_id = r.id
        WHERE mpw.plan_id = ?
        ORDER BY mpw.week_number, mpm.day_of_week, mpm.meal_type DESC
        """,
        (plan_id,),
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_all_ingredients_for_plan(plan_id: int) -> list:
    """Return all ingredients needed for the entire 12-week plan, aggregated."""
    conn = get_conn()
    rows = conn.execute(
        """
        SELECT 
            i.id,
            i.name,
            i.category,
            i.store_section,
            SUM(ri.quantity) as total_quantity,
            MAX(ri.unit) as unit,
            COUNT(DISTINCT mpm.id) as used_in_meals,
            COUNT(DISTINCT mpw.week_number) as used_in_weeks
        FROM meal_plan_meals mpm
        JOIN meal_plan_weeks mpw ON mpm.week_id = mpw.id
        JOIN recipe_ingredients ri ON mpm.recipe_id = ri.recipe_id
        JOIN ingredients i ON ri.ingredient_id = i.id
        WHERE mpw.plan_id = ?
        GROUP BY i.id, i.name, i.category, i.store_section, ri.unit
        ORDER BY i.store_section, i.name
        """,
        (plan_id,),
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_weekly_ingredients(plan_id: int, week_number: int) -> list:
    """Return ingredients needed for a specific week."""
    conn = get_conn()
    rows = conn.execute(
        """
        SELECT 
            i.id,
            i.name,
            i.category,
            i.store_section,
            SUM(ri.quantity) as total_quantity,
            MAX(ri.unit) as unit,
            COUNT(DISTINCT mpm.id) as used_in_meals
        FROM meal_plan_meals mpm
        JOIN meal_plan_weeks mpw ON mpm.week_id = mpw.id
        JOIN recipe_ingredients ri ON mpm.recipe_id = ri.recipe_id
        JOIN ingredients i ON ri.ingredient_id = i.id
        WHERE mpw.plan_id = ? AND mpw.week_number = ?
        GROUP BY i.id, i.name, i.category, i.store_section, ri.unit
        ORDER BY i.store_section, i.name
        """,
        (plan_id, week_number),
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_recipe_details(recipe_id: int) -> dict:
    """Get full recipe with ingredient list."""
    conn = get_conn()
    recipe = conn.execute("SELECT * FROM recipes WHERE id = ?", (recipe_id,)).fetchone()
    if not recipe:
        conn.close()
        return None
    
    ingredients = conn.execute(
        """
        SELECT ri.quantity, ri.unit, ri.prep_note, ri.optional, i.name, i.category
        FROM recipe_ingredients ri
        JOIN ingredients i ON ri.ingredient_id = i.id
        WHERE ri.recipe_id = ?
        ORDER BY i.category, i.name
        """,
        (recipe_id,),
    ).fetchall()
    conn.close()
    return {
        **dict(recipe),
        "ingredients": [dict(i) for i in ingredients]
    }


def export_meal_plan_csv(output_path: str, plan_id: int):
    """Export the 12-week meal plan as CSV."""
    meals = get_full_plan(plan_id)
    DAYS = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
    
    with open(output_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow([
            "Week", "Day", "Meal_Type", "Recipe_ID", "Recipe_Name", 
            "Cuisine", "Prep_Time_Min", "Cook_Time_Min", "Servings", "Difficulty", "Tags"
        ])
        for m in meals:
            writer.writerow([
                m["week_number"],
                DAYS[m["day_of_week"]],
                m["meal_type"],
                m["recipe_id"],
                m["recipe_name"],
                m["cuisine"],
                m["prep_time"],
                m["cook_time"],
                m["servings"],
                m["difficulty"],
                m["tags"],
            ])
    print(f"Exported meal plan: {output_path}")


def export_ingredients_csv(output_path: str, plan_id: int):
    """Export aggregated ingredient list for the whole plan."""
    ingredients = get_all_ingredients_for_plan(plan_id)
    
    with open(output_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow([
            "Ingredient_ID", "Name", "Category", "Store_Section",
            "Total_Quantity", "Unit", "Used_In_Meals", "Used_In_Weeks"
        ])
        for ing in ingredients:
            writer.writerow([
                ing["id"],
                ing["name"],
                ing["category"],
                ing["store_section"],
                round(ing["total_quantity"], 1),
                ing["unit"],
                ing["used_in_meals"],
                ing["used_in_weeks"],
            ])
    print(f"Exported ingredients: {output_path}")


def export_weekly_shopping_csv(output_dir: str, plan_id: int):
    """Export a separate CSV for each week's shopping list."""
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    
    for week in range(1, 13):
        ingredients = get_weekly_ingredients(plan_id, week)
        path = output_dir / f"week_{week:02d}_shopping.csv"
        
        with open(path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(["Ingredient", "Category", "Store_Section", "Quantity", "Unit", "Needed_For_N_Meals"])
            for ing in ingredients:
                writer.writerow([
                    ing["name"],
                    ing["category"],
                    ing["store_section"],
                    round(ing["total_quantity"], 1),
                    ing["unit"],
                    ing["used_in_meals"],
                ])
        print(f"  Week {week}: {path}")


def export_recipe_book_csv(output_path: str, plan_id: int):
    """Export all recipes in the plan as a recipe book CSV."""
    conn = get_conn()
    recipe_ids = conn.execute(
        """
        SELECT DISTINCT mpm.recipe_id
        FROM meal_plan_meals mpm
        JOIN meal_plan_weeks mpw ON mpm.week_id = mpw.id
        WHERE mpw.plan_id = ?
        ORDER BY mpm.recipe_id
        """,
        (plan_id,),
    ).fetchall()
    conn.close()
    
    with open(output_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow([
            "Recipe_ID", "Name", "Type", "Cuisine", "Difficulty",
            "Prep_Time", "Cook_Time", "Servings", "Tags",
            "Ingredients_List"
        ])
        for rid in recipe_ids:
            r = get_recipe_details(rid[0])
            if not r:
                continue
            ing_list = "; ".join([
                f"{i['quantity']}{i['unit']} {i['name']}{(' (' + i['prep_note'] + ')') if i['prep_note'] else ''}"
                for i in r["ingredients"]
            ])
            writer.writerow([
                r["id"],
                r["name"],
                r["type"],
                r["cuisine"],
                r["difficulty"],
                r["prep_time"],
                r["cook_time"],
                r["servings"],
                r["tags"],
                ing_list,
            ])
    print(f"Exported recipe book: {output_path}")


def export_xlsx(output_path: str, plan_id: int):
    """Export full data to multi-sheet Excel file."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("openpyxl not installed. Installing...")
        import subprocess
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    wb = openpyxl.Workbook()
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    def write_sheet(ws, headers, rows, sheet_title):
        ws.title = sheet_title
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        for row in rows:
            ws.append(row)
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Sheet 1: Meal Plan
    meals = get_full_plan(plan_id)
    DAYS = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
    headers1 = ["Week", "Day", "Meal", "Recipe", "Cuisine", "Prep (min)", "Cook (min)", "Servings", "Difficulty", "Tags"]
    rows1 = [
        [m["week_number"], DAYS[m["day_of_week"]], m["meal_type"].capitalize(),
         m["recipe_name"], m["cuisine"], m["prep_time"], m["cook_time"],
         m["servings"], m["difficulty"], m["tags"]]
        for m in meals
    ]
    write_sheet(wb.active, headers1, rows1, "Meal Plan")
    
    # Sheet 2: All Ingredients
    ingredients = get_all_ingredients_for_plan(plan_id)
    headers2 = ["Ingredient", "Category", "Store Section", "Total Quantity", "Unit", "Used in Meals", "Used in Weeks"]
    rows2 = [
        [ing["name"], ing["category"], ing["store_section"],
         round(ing["total_quantity"], 1), ing["unit"],
         ing["used_in_meals"], ing["used_in_weeks"]]
        for ing in ingredients
    ]
    ws2 = wb.create_sheet(title="All Ingredients")
    write_sheet(ws2, headers2, rows2, "All Ingredients")
    
    # Sheet 3: Weekly Shopping (concatenated)
    ws3 = wb.create_sheet(title="Weekly Shopping")
    ws3.append(["Week", "Ingredient", "Category", "Store Section", "Quantity", "Unit"])
    for cell in ws3[1]:
        cell.font = header_font
        cell.fill = header_fill
    for week in range(1, 13):
        week_ings = get_weekly_ingredients(plan_id, week)
        for ing in week_ings:
            ws3.append([week, ing["name"], ing["category"], ing["store_section"],
                        round(ing["total_quantity"], 1), ing["unit"]])
    
    # Sheet 4: Recipes
    conn = get_conn()
    recipe_ids = conn.execute(
        """SELECT DISTINCT mpm.recipe_id FROM meal_plan_meals mpm
           JOIN meal_plan_weeks mpw ON mpm.week_id = mpw.id
           WHERE mpw.plan_id = ? ORDER BY mpm.recipe_id""", (plan_id,)
    ).fetchall()
    conn.close()
    
    ws4 = wb.create_sheet(title="Recipes")
    ws4.append(["ID", "Name", "Type", "Cuisine", "Difficulty", "Prep", "Cook", "Servings", "Tags", "Ingredients"])
    for cell in ws4[1]:
        cell.font = header_font
        cell.fill = header_fill
    for rid in recipe_ids:
        r = get_recipe_details(rid[0])
        if r:
            ing_list = "; ".join([
                f"{i['quantity']}{i['unit']} {i['name']}{(' (' + i['prep_note'] + ')') if i['prep_note'] else ''}"
                for i in r["ingredients"]
            ])
            ws4.append([r["id"], r["name"], r["type"], r["cuisine"], r["difficulty"],
                        r["prep_time"], r["cook_time"], r["servings"], r["tags"], ing_list])
    
    wb.save(output_path)
    print(f"Exported Excel workbook: {output_path}")
    print(f"  Sheets: Meal Plan, All Ingredients, Weekly Shopping, Recipes")


def main():
    parser = argparse.ArgumentParser(description="Export vdM meal plan to spreadsheet")
    parser.add_argument("--format", choices=["csv", "xlsx"], default="xlsx",
                        help="Output format (default: xlsx)")
    parser.add_argument("--output", default=None,
                        help="Output file path (default: vdM_MealPlan.xlsx or ./exports/)")
    parser.add_argument("--output-dir", default="./exports",
                        help="Directory for CSV exports (default: ./exports)")
    args = parser.parse_args()
    
    plan = get_active_plan()
    if not plan:
        print("No active meal plan found.")
        sys.exit(1)
    plan_id = plan["id"]
    print(f"Exporting plan: {plan['name']} (ID: {plan_id})")
    
    if args.format == "csv":
        out_dir = Path(args.output_dir)
        out_dir.mkdir(parents=True, exist_ok=True)
        export_meal_plan_csv(str(out_dir / "meal_plan.csv"), plan_id)
        export_ingredients_csv(str(out_dir / "all_ingredients.csv"), plan_id)
        export_weekly_shopping_csv(str(out_dir / "weekly"), plan_id)
        export_recipe_book_csv(str(out_dir / "recipe_book.csv"), plan_id)
        print(f"\nAll CSVs exported to: {out_dir.absolute()}")
    else:
        output = args.output or "vdM_MealPlan.xlsx"
        export_xlsx(output, plan_id)


if __name__ == "__main__":
    main()
